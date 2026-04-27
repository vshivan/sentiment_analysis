"""
Senti ERP - Complete Backend API
Sprint 1: Auth + SQLite + Validation + Rate Limiting + Audit Log
Sprint 2: Pagination + Date stamps + RBAC + Excel export
"""
import os, csv, io, json, re, logging
from datetime import datetime, timezone, timedelta
from functools import wraps

from flask import Flask, jsonify, request, Response, g
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from flask_jwt_extended import (
    JWTManager, create_access_token, jwt_required,
    get_jwt_identity, verify_jwt_in_request
)
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from sentiment import analyze_sentiment, get_sentiment_stats, extract_keywords
from scraper import scrape_reviews

load_dotenv()

# ── App setup ────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.config["SECRET_KEY"]                  = os.getenv("SECRET_KEY", "dev-secret")
app.config["JWT_SECRET_KEY"]              = os.getenv("JWT_SECRET_KEY", "dev-jwt")
app.config["JWT_ACCESS_TOKEN_EXPIRES"]    = timedelta(hours=8)
app.config["SQLALCHEMY_DATABASE_URI"]     = os.getenv("DATABASE_URL", "sqlite:///senti.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

CORS(app, resources={r"/*": {"origins": "*"}})

db      = SQLAlchemy(app)
jwt     = JWTManager(app)
limiter = Limiter(key_func=get_remote_address, app=app,
                  default_limits=["500 per hour", "100 per minute"])

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("senti")

# ── Models ────────────────────────────────────────────────────────────────────
class User(db.Model):
    __tablename__ = "users"
    id         = db.Column(db.Integer, primary_key=True)
    username   = db.Column(db.String(80),  unique=True, nullable=False)
    email      = db.Column(db.String(120), unique=True, nullable=False)
    password   = db.Column(db.String(256), nullable=False)
    role       = db.Column(db.String(20),  default="viewer")
    is_active  = db.Column(db.Boolean,     default=True)
    created_at = db.Column(db.DateTime,    default=lambda: datetime.now(timezone.utc))
    last_login = db.Column(db.DateTime,    nullable=True)

    def to_dict(self):
        return {"id": self.id, "username": self.username, "email": self.email,
                "role": self.role, "is_active": self.is_active,
                "created_at": self.created_at.isoformat(),
                "last_login": self.last_login.isoformat() if self.last_login else None}


class Product(db.Model):
    __tablename__ = "products"
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(100), unique=True, nullable=False)
    category   = db.Column(db.String(50),  default="General")
    icon       = db.Column(db.String(10),  default="📦")
    is_active  = db.Column(db.Boolean,     default=True)
    created_at = db.Column(db.DateTime,    default=lambda: datetime.now(timezone.utc))
    reviews    = db.relationship("Review", backref="product", lazy=True,
                                 cascade="all, delete-orphan")

    def to_dict(self):
        return {"id": self.id, "name": self.name, "category": self.category,
                "icon": self.icon, "is_active": self.is_active,
                "created_at": self.created_at.isoformat(),
                "review_count": len(self.reviews)}


class Review(db.Model):
    __tablename__ = "reviews"
    id         = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey("products.id"), nullable=False)
    text       = db.Column(db.Text,    nullable=False)
    sentiment  = db.Column(db.String(20), nullable=True)
    score      = db.Column(db.Float,      nullable=True)
    language   = db.Column(db.String(20), default="mixed")
    source     = db.Column(db.String(50), default="manual")
    tags       = db.Column(db.String(200), default="")
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    created_by = db.Column(db.String(80), default="system")

    def to_dict(self):
        return {"id": self.id, "product_id": self.product_id,
                "product_name": self.product.name if self.product else "",
                "text": self.text, "sentiment": self.sentiment,
                "score": self.score, "language": self.language,
                "source": self.source, "tags": self.tags,
                "created_at": self.created_at.isoformat(),
                "created_by": self.created_by}


class AuditLog(db.Model):
    __tablename__ = "audit_logs"
    id         = db.Column(db.Integer, primary_key=True)
    user       = db.Column(db.String(80),  nullable=False)
    action     = db.Column(db.String(50),  nullable=False)
    resource   = db.Column(db.String(100), nullable=False)
    detail     = db.Column(db.Text,        nullable=True)
    ip_address = db.Column(db.String(45),  nullable=True)
    created_at = db.Column(db.DateTime,    default=lambda: datetime.now(timezone.utc))

    def to_dict(self):
        return {"id": self.id, "user": self.user, "action": self.action,
                "resource": self.resource, "detail": self.detail,
                "ip_address": self.ip_address,
                "created_at": self.created_at.isoformat()}


class Notification(db.Model):
    __tablename__ = "notifications"
    id         = db.Column(db.Integer, primary_key=True)
    user       = db.Column(db.String(80),  nullable=False)
    title      = db.Column(db.String(120), nullable=False)
    message    = db.Column(db.Text,        nullable=False)
    type       = db.Column(db.String(20),  default="info")
    is_read    = db.Column(db.Boolean,     default=False)
    created_at = db.Column(db.DateTime,    default=lambda: datetime.now(timezone.utc))

    def to_dict(self):
        return {"id": self.id, "user": self.user, "title": self.title,
                "message": self.message, "type": self.type,
                "is_read": self.is_read,
                "created_at": self.created_at.isoformat()}


# ── Helpers ───────────────────────────────────────────────────────────────────
PRODUCT_ICONS = {
    "Adidas": "👟", "Zara": "👗", "Dell": "💻", "Supra": "👠",
    "iPhone": "📱", "Lenskart": "👓", "Lloyd AC": "❄️", "Titan Watch": "⌚"
}

SEED_DATA = {
    "Adidas":       ["The shoes are super comfortable and perfect for running. Totally worth it.",
                     "Quality achhi hai but price thoda zyada lagta hai.",
                     "I love the design, very stylish and durable.",
                     "Ek mahine ke baad sole thoda ghis gaya, disappointed.",
                     "Best sports brand experience so far, bahut badhiya."],
    "Zara":         ["Clothes look premium but stitching could be better.",
                     "Zara ka collection trendy hota hai, mujhe pasand aaya.",
                     "Fabric quality is average, not worth the high price.",
                     "Amazing designs, feels very fashionable.",
                     "Kapde ache hain lekin size fitting issue hai."],
    "Dell":         ["Laptop performance is smooth and reliable.",
                     "Battery backup utna achha nahi hai.",
                     "Dell ka build quality solid hai.",
                     "Got heating issues after some months.",
                     "Perfect for students and office work."],
    "Supra":        ["Shoes look cool but comfort could improve.",
                     "Design mast hai, youth ke liye perfect.",
                     "Material quality is not very durable.",
                     "Affordable and stylish option.",
                     "Thoda heavy feel hota hai pehenne mein."],
    "iPhone":       ["Camera quality is outstanding, best in class.",
                     "Bahut mehenga hai, sabke liye nahi hai.",
                     "Performance is super smooth, no lag at all.",
                     "Battery life improve ho sakti thi.",
                     "Premium feel and great ecosystem."],
    "Lenskart":     ["Affordable glasses and good service.",
                     "Frame quality utni strong nahi lagi.",
                     "Delivery fast thi aur fitting perfect.",
                     "Customer service could be better.",
                     "Great variety of stylish frames."],
    "Lloyd AC":     ["Cooling is very effective even in peak summer.",
                     "Installation service late tha.",
                     "Energy efficient and silent operation.",
                     "Remote kabhi kabhi properly work nahi karta.",
                     "Value for money product."],
    "Titan Watch":  ["Elegant design and premium feel.",
                     "Thoda expensive hai but quality achhi hai.",
                     "Strap quality could be better.",
                     "Perfect gift option.",
                     "Time accuracy aur durability top notch hai."]
}

CATEGORIES = {
    "Adidas": "Footwear", "Zara": "Fashion", "Dell": "Electronics",
    "Supra": "Footwear", "iPhone": "Electronics", "Lenskart": "Eyewear",
    "Lloyd AC": "Appliances", "Titan Watch": "Accessories"
}

def detect_language(text):
    hindi_chars = len(re.findall(r"[\u0900-\u097F]", text))
    hinglish    = any(w in text.lower() for w in ["hai","ka","ki","ke","se","mein","nahi","aur","bahut","achha","achhi","tha","thi"])
    if hindi_chars > 3:   return "hindi"
    if hinglish:          return "hinglish"
    return "english"

def audit(action, resource, detail=None):
    try:
        user = "system"
        try:
            verify_jwt_in_request(optional=True)
            u = get_jwt_identity()
            if u: user = u
        except Exception:
            pass
        log = AuditLog(user=user, action=action, resource=resource,
                       detail=detail, ip_address=request.remote_addr)
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        logger.error(f"Audit log failed: {e}")

def notify(user, title, message, ntype="info"):
    try:
        n = Notification(user=user, title=title, message=message, type=ntype)
        db.session.add(n)
        db.session.commit()
    except Exception as e:
        logger.error(f"Notification failed: {e}")

def roles_required(*roles):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            verify_jwt_in_request()
            identity = get_jwt_identity()
            user = User.query.filter_by(username=identity).first()
            if not user or user.role not in roles:
                return jsonify({"error": "Insufficient permissions"}), 403
            if not user.is_active:
                return jsonify({"error": "Account deactivated"}), 403
            return fn(*args, **kwargs)
        return wrapper
    return decorator

def paginate(query, page, per_page=20):
    total   = query.count()
    items   = query.offset((page - 1) * per_page).limit(per_page).all()
    return {
        "items": items, "total": total, "page": page,
        "per_page": per_page, "pages": max(1, (total + per_page - 1) // per_page),
        "has_next": page * per_page < total,
        "has_prev": page > 1
    }


# ── DB Init + Seed ────────────────────────────────────────────────────────────
def init_db():
    with app.app_context():
        db.create_all()
        # Seed admin user
        if not User.query.filter_by(username="admin").first():
            admin = User(username="admin", email="admin@senti.io",
                         password=generate_password_hash("admin123"),
                         role="admin")
            analyst = User(username="analyst", email="analyst@senti.io",
                           password=generate_password_hash("analyst123"),
                           role="analyst")
            viewer = User(username="viewer", email="viewer@senti.io",
                          password=generate_password_hash("viewer123"),
                          role="viewer")
            db.session.add_all([admin, analyst, viewer])
            db.session.commit()
            logger.info("Seeded default users")

        # Seed products + reviews
        for pname, reviews in SEED_DATA.items():
            if not Product.query.filter_by(name=pname).first():
                p = Product(name=pname,
                            category=CATEGORIES.get(pname, "General"),
                            icon=PRODUCT_ICONS.get(pname, "📦"))
                db.session.add(p)
                db.session.flush()
                for text in reviews:
                    sentiment, score = analyze_sentiment(text)
                    r = Review(product_id=p.id, text=text,
                               sentiment=sentiment, score=round(score, 4),
                               language=detect_language(text),
                               source="seed", created_by="system")
                    db.session.add(r)
        db.session.commit()
        logger.info("Database initialized and seeded")

# ── AUTH ROUTES ───────────────────────────────────────────────────────────────
@app.route("/auth/login", methods=["POST"])
@limiter.limit("10 per minute")
def login():
    body = request.get_json() or {}
    username = body.get("username", "").strip()
    password = body.get("password", "")
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    user = User.query.filter_by(username=username).first()
    if not user or not check_password_hash(user.password, password):
        return jsonify({"error": "Invalid credentials"}), 401
    if not user.is_active:
        return jsonify({"error": "Account deactivated"}), 403
    user.last_login = datetime.now(timezone.utc)
    db.session.commit()
    token = create_access_token(identity=username)
    audit("LOGIN", f"User: {username}")
    return jsonify({"token": token, "user": user.to_dict()})

@app.route("/auth/me", methods=["GET"])
@jwt_required()
def me():
    identity = get_jwt_identity()
    user = User.query.filter_by(username=identity).first()
    if not user:
        return jsonify({"error": "User not found"}), 404
    return jsonify({"user": user.to_dict()})

@app.route("/auth/register", methods=["POST"])
@jwt_required()
@roles_required("admin")
def register():
    body = request.get_json() or {}
    username = body.get("username", "").strip()
    email    = body.get("email", "").strip()
    password = body.get("password", "")
    role     = body.get("role", "viewer")
    if not all([username, email, password]):
        return jsonify({"error": "username, email, password required"}), 400
    if role not in ("admin", "analyst", "viewer"):
        return jsonify({"error": "Invalid role"}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({"error": "Username already exists"}), 409
    user = User(username=username, email=email,
                password=generate_password_hash(password), role=role)
    db.session.add(user)
    db.session.commit()
    audit("CREATE", f"User: {username}", f"role={role}")
    return jsonify({"message": "User created", "user": user.to_dict()}), 201

@app.route("/auth/users", methods=["GET"])
@jwt_required()
@roles_required("admin")
def list_users():
    users = User.query.all()
    return jsonify({"users": [u.to_dict() for u in users]})

@app.route("/auth/users/<int:uid>", methods=["PATCH"])
@jwt_required()
@roles_required("admin")
def update_user(uid):
    user = User.query.get_or_404(uid)
    body = request.get_json() or {}
    if "role" in body and body["role"] in ("admin", "analyst", "viewer"):
        user.role = body["role"]
    if "is_active" in body:
        user.is_active = bool(body["is_active"])
    db.session.commit()
    audit("UPDATE", f"User: {user.username}", str(body))
    return jsonify({"user": user.to_dict()})


# ── PRODUCTS ──────────────────────────────────────────────────────────────────
@app.route("/products", methods=["GET"])
@jwt_required()
def get_products():
    products = Product.query.filter_by(is_active=True).all()
    return jsonify({"products": [p.name for p in products],
                    "details":  [p.to_dict() for p in products]})

@app.route("/products", methods=["POST"])
@jwt_required()
@roles_required("admin", "analyst")
def create_product():
    body = request.get_json() or {}
    name = body.get("name", "").strip()
    if not name:
        return jsonify({"error": "Product name required"}), 400
    if Product.query.filter_by(name=name).first():
        return jsonify({"error": "Product already exists"}), 409
    p = Product(name=name, category=body.get("category","General"),
                icon=body.get("icon","📦"))
    db.session.add(p)
    db.session.commit()
    audit("CREATE", f"Product: {name}")
    return jsonify({"product": p.to_dict()}), 201

@app.route("/products/<int:pid>", methods=["DELETE"])
@jwt_required()
@roles_required("admin")
def delete_product(pid):
    p = Product.query.get_or_404(pid)
    p.is_active = False
    db.session.commit()
    audit("DELETE", f"Product: {p.name}")
    return jsonify({"message": f"Product {p.name} deactivated"})

# ── REVIEWS ───────────────────────────────────────────────────────────────────
@app.route("/reviews/<product_name>", methods=["GET"])
@jwt_required()
def get_reviews(product_name):
    p = Product.query.filter_by(name=product_name, is_active=True).first()
    if not p:
        return jsonify({"error": "Product not found"}), 404

    page     = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 20))
    sentiment_filter = request.args.get("sentiment")
    search   = request.args.get("search", "").strip()
    sort     = request.args.get("sort", "newest")

    q = Review.query.filter_by(product_id=p.id)
    if sentiment_filter:
        q = q.filter_by(sentiment=sentiment_filter)
    if search:
        q = q.filter(Review.text.ilike(f"%{search}%"))
    if sort == "newest":
        q = q.order_by(Review.created_at.desc())
    elif sort == "oldest":
        q = q.order_by(Review.created_at.asc())
    elif sort == "score-desc":
        q = q.order_by(Review.score.desc())
    elif sort == "score-asc":
        q = q.order_by(Review.score.asc())

    result = paginate(q, page, per_page)
    return jsonify({
        "product": product_name,
        "reviews": [r.to_dict() for r in result["items"]],
        "pagination": {k: v for k, v in result.items() if k != "items"}
    })

@app.route("/reviews/<product_name>/add", methods=["POST"])
@jwt_required()
@roles_required("admin", "analyst")
def add_review(product_name):
    p = Product.query.filter_by(name=product_name, is_active=True).first()
    if not p:
        return jsonify({"error": "Product not found"}), 404
    body = request.get_json() or {}
    text = body.get("text", "").strip()
    if not text or len(text) < 5:
        return jsonify({"error": "Review text too short (min 5 chars)"}), 400
    if len(text) > 2000:
        return jsonify({"error": "Review text too long (max 2000 chars)"}), 400

    sentiment, score = analyze_sentiment(text)
    identity = get_jwt_identity()
    r = Review(product_id=p.id, text=text, sentiment=sentiment,
               score=round(score, 4), language=detect_language(text),
               source=body.get("source", "manual"),
               tags=body.get("tags", ""), created_by=identity)
    db.session.add(r)
    db.session.commit()
    audit("CREATE", f"Review #{r.id}", f"product={product_name} sentiment={sentiment}")

    # Auto-alert if negative
    if sentiment == "Negative":
        neg_count = Review.query.filter_by(product_id=p.id, sentiment="Negative").count()
        total     = Review.query.filter_by(product_id=p.id).count()
        if total > 0 and (neg_count / total) >= 0.4:
            notify("admin", f"Alert: {product_name}",
                   f"Negative sentiment reached {round(neg_count/total*100)}%", "alert")

    return jsonify({"review": r.to_dict()}), 201

@app.route("/reviews/<product_name>/<int:rid>", methods=["DELETE"])
@jwt_required()
@roles_required("admin", "analyst")
def delete_review(product_name, rid):
    r = Review.query.get_or_404(rid)
    audit("DELETE", f"Review #{rid}", f"text={r.text[:60]}")
    db.session.delete(r)
    db.session.commit()
    return jsonify({"message": "Review deleted"})

@app.route("/reviews/<product_name>/<int:rid>", methods=["PATCH"])
@jwt_required()
@roles_required("admin", "analyst")
def update_review(product_name, rid):
    r = Review.query.get_or_404(rid)
    body = request.get_json() or {}
    if "tags" in body:
        r.tags = body["tags"]
    if "text" in body:
        r.text = body["text"].strip()
        r.sentiment, score = analyze_sentiment(r.text)
        r.score = round(score, 4)
        r.language = detect_language(r.text)
    db.session.commit()
    audit("UPDATE", f"Review #{rid}", str(body))
    return jsonify({"review": r.to_dict()})

# ── GLOBAL SEARCH ─────────────────────────────────────────────────────────────
@app.route("/search", methods=["GET"])
@jwt_required()
def global_search():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify({"error": "Query too short"}), 400
    page     = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 20))

    query = Review.query.filter(Review.text.ilike(f"%{q}%"))
    result = paginate(query, page, per_page)
    return jsonify({
        "query": q,
        "reviews": [r.to_dict() for r in result["items"]],
        "pagination": {k: v for k, v in result.items() if k != "items"}
    })


# ── STATS + ANALYTICS ─────────────────────────────────────────────────────────
@app.route("/stats", methods=["GET"])
@jwt_required()
def stats():
    all_stats = {}
    total_reviews = 0
    overall = {"Positive": 0, "Negative": 0, "Neutral": 0}

    for p in Product.query.filter_by(is_active=True).all():
        reviews = Review.query.filter_by(product_id=p.id).all()
        texts   = [r.text for r in reviews]
        s       = get_sentiment_stats(texts) if texts else {
            "total":0,"Positive":0,"Negative":0,"Neutral":0,
            "positive_pct":0,"negative_pct":0,"neutral_pct":0,"avg_score":0
        }
        all_stats[p.name] = s
        total_reviews += s["total"]
        for k in overall:
            overall[k] += s[k]

    return jsonify({
        "total_reviews": total_reviews,
        "overall_sentiment": overall,
        "product_stats": all_stats,
        "positive_pct": round(overall["Positive"]/total_reviews*100,1) if total_reviews else 0,
        "negative_pct": round(overall["Negative"]/total_reviews*100,1) if total_reviews else 0,
        "neutral_pct":  round(overall["Neutral"] /total_reviews*100,1) if total_reviews else 0,
    })

@app.route("/keywords/<product_name>", methods=["GET"])
@jwt_required()
def keywords(product_name):
    p = Product.query.filter_by(name=product_name, is_active=True).first()
    if not p:
        return jsonify({"error": "Product not found"}), 404
    texts = [r.text for r in Review.query.filter_by(product_id=p.id).all()]
    return jsonify({"product": product_name, "keywords": extract_keywords(texts)})

@app.route("/analyze", methods=["POST"])
@jwt_required()
def analyze():
    body = request.get_json() or {}
    text = body.get("text", "").strip()
    if not text:
        return jsonify({"error": "Missing text"}), 400
    sentiment, score = analyze_sentiment(text)
    return jsonify({"text": text, "sentiment": sentiment,
                    "score": round(score, 4), "language": detect_language(text)})

@app.route("/leaderboard", methods=["GET"])
@jwt_required()
def leaderboard():
    board = []
    for p in Product.query.filter_by(is_active=True).all():
        reviews  = Review.query.filter_by(product_id=p.id).all()
        texts    = [r.text for r in reviews]
        s        = get_sentiment_stats(texts) if texts else {"total":0,"Positive":0,"Negative":0,"Neutral":0,"positive_pct":0,"negative_pct":0,"neutral_pct":0,"avg_score":0}
        enriched = [{"id":r.id,"text":r.text,"sentiment":r.sentiment,"score":r.score} for r in reviews]
        board.append({
            "product": p.name, "icon": p.icon, "category": p.category,
            "total": s["total"], "positive": s["Positive"],
            "negative": s["Negative"], "neutral": s["Neutral"],
            "positive_pct": s["positive_pct"], "negative_pct": s["negative_pct"],
            "neutral_pct": s["neutral_pct"], "avg_score": s["avg_score"],
            "top_review":   max(enriched, key=lambda r: r["score"])["text"] if enriched else "",
            "worst_review": min(enriched, key=lambda r: r["score"])["text"] if enriched else "",
        })
    board.sort(key=lambda x: x["positive_pct"], reverse=True)
    for i, item in enumerate(board):
        item["rank"] = i + 1
    return jsonify({"leaderboard": board})

@app.route("/alerts", methods=["GET"])
@jwt_required()
def alerts():
    threshold = float(request.args.get("threshold", 40))
    result = []
    for p in Product.query.filter_by(is_active=True).all():
        reviews = Review.query.filter_by(product_id=p.id).all()
        texts   = [r.text for r in reviews]
        s       = get_sentiment_stats(texts) if texts else {"negative_pct":0,"Negative":0,"total":0,"avg_score":0}
        if s["negative_pct"] >= threshold:
            result.append({
                "product": p.name, "icon": p.icon,
                "negative_pct": s["negative_pct"],
                "negative": s["Negative"], "total": s["total"],
                "avg_score": s["avg_score"],
                "severity": "high" if s["negative_pct"] >= 60 else "medium"
            })
    result.sort(key=lambda x: x["negative_pct"], reverse=True)
    return jsonify({"alerts": result, "threshold": threshold})

@app.route("/compare", methods=["GET"])
@jwt_required()
def compare():
    names = request.args.get("products", "")
    product_list = [n.strip() for n in names.split(",") if n.strip()]
    if len(product_list) < 2:
        return jsonify({"error": "Provide at least 2 products"}), 400
    result = {}
    for name in product_list:
        p = Product.query.filter_by(name=name, is_active=True).first()
        if not p:
            return jsonify({"error": f"Product not found: {name}"}), 404
        reviews = Review.query.filter_by(product_id=p.id).all()
        texts   = [r.text for r in reviews]
        result[name] = {
            "stats":    get_sentiment_stats(texts) if texts else {},
            "reviews":  [r.to_dict() for r in reviews],
            "keywords": extract_keywords(texts, top_n=15)
        }
    return jsonify({"comparison": result})


# ── EXPORT ────────────────────────────────────────────────────────────────────
@app.route("/export/<product_name>/csv", methods=["GET"])
@jwt_required()
def export_csv(product_name):
    p = Product.query.filter_by(name=product_name, is_active=True).first()
    if not p:
        return jsonify({"error": "Product not found"}), 404
    reviews = Review.query.filter_by(product_id=p.id).all()
    output  = io.StringIO()
    writer  = csv.DictWriter(output, fieldnames=["id","product","review","sentiment","score","language","source","created_at","created_by"])
    writer.writeheader()
    for r in reviews:
        writer.writerow({"id":r.id,"product":product_name,"review":r.text,
                         "sentiment":r.sentiment,"score":r.score,
                         "language":r.language,"source":r.source,
                         "created_at":r.created_at.isoformat(),"created_by":r.created_by})
    audit("EXPORT", f"Product: {product_name}", "format=csv")
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={product_name}_reviews.csv"})

@app.route("/export/<product_name>/json", methods=["GET"])
@jwt_required()
def export_json(product_name):
    p = Product.query.filter_by(name=product_name, is_active=True).first()
    if not p:
        return jsonify({"error": "Product not found"}), 404
    reviews = Review.query.filter_by(product_id=p.id).all()
    data    = json.dumps({"product": product_name, "reviews": [r.to_dict() for r in reviews]},
                         indent=2, ensure_ascii=False)
    audit("EXPORT", f"Product: {product_name}", "format=json")
    return Response(data, mimetype="application/json",
                    headers={"Content-Disposition": f"attachment; filename={product_name}_reviews.json"})

@app.route("/export/<product_name>/xlsx", methods=["GET"])
@jwt_required()
def export_xlsx(product_name):
    p = Product.query.filter_by(name=product_name, is_active=True).first()
    if not p:
        return jsonify({"error": "Product not found"}), 404
    reviews = Review.query.filter_by(product_id=p.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = product_name[:31]

    # Header styling
    header_fill = PatternFill("solid", fgColor="1a1d2e")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["#","Product","Review","Sentiment","Score","Language","Source","Date","Added By"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Sentiment colors
    sent_colors = {"Positive": "d4edda", "Negative": "f8d7da", "Neutral": "fff3cd"}
    for row, r in enumerate(reviews, 2):
        ws.cell(row=row, column=1, value=r.id)
        ws.cell(row=row, column=2, value=product_name)
        ws.cell(row=row, column=3, value=r.text)
        sent_cell = ws.cell(row=row, column=4, value=r.sentiment)
        sent_cell.fill = PatternFill("solid", fgColor=sent_colors.get(r.sentiment, "ffffff"))
        ws.cell(row=row, column=5, value=r.score)
        ws.cell(row=row, column=6, value=r.language)
        ws.cell(row=row, column=7, value=r.source)
        ws.cell(row=row, column=8, value=r.created_at.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=9, value=r.created_by)

    # Column widths
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 12
    for col in ["A","B","E","F","G","H","I"]:
        ws.column_dimensions[col].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    audit("EXPORT", f"Product: {product_name}", "format=xlsx")
    return Response(output.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={product_name}_reviews.xlsx"})

@app.route("/export/all/csv", methods=["GET"])
@jwt_required()
def export_all_csv():
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["id","product","review","sentiment","score","language","source","created_at"])
    writer.writeheader()
    for p in Product.query.filter_by(is_active=True).all():
        for r in Review.query.filter_by(product_id=p.id).all():
            writer.writerow({"id":r.id,"product":p.name,"review":r.text,
                             "sentiment":r.sentiment,"score":r.score,
                             "language":r.language,"source":r.source,
                             "created_at":r.created_at.isoformat()})
    audit("EXPORT", "All Products", "format=csv")
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=senti_all_reviews.csv"})

# ── BULK IMPORT ───────────────────────────────────────────────────────────────
@app.route("/import/<product_name>", methods=["POST"])
@jwt_required()
@roles_required("admin", "analyst")
def bulk_import(product_name):
    p = Product.query.filter_by(name=product_name, is_active=True).first()
    if not p:
        return jsonify({"error": "Product not found"}), 404
    body  = request.get_json() or {}
    texts = body.get("reviews", [])
    if not texts:
        return jsonify({"error": "No reviews provided"}), 400

    identity = get_jwt_identity()
    added = 0
    for text in texts:
        if isinstance(text, str) and text.strip() and len(text.strip()) >= 5:
            sentiment, score = analyze_sentiment(text.strip())
            r = Review(product_id=p.id, text=text.strip(),
                       sentiment=sentiment, score=round(score,4),
                       language=detect_language(text),
                       source="import", created_by=identity)
            db.session.add(r)
            added += 1
    db.session.commit()
    audit("IMPORT", f"Product: {product_name}", f"added={added}")
    return jsonify({"message": f"Imported {added} reviews", "added": added})

# ── SCRAPE ────────────────────────────────────────────────────────────────────
@app.route("/scrape", methods=["POST"])
@jwt_required()
@limiter.limit("10 per minute")
def scrape():
    body = request.get_json() or {}
    url  = body.get("url", "").strip()
    if not url:
        return jsonify({"error": "URL required"}), 400
    scraped = scrape_reviews(url)
    if not scraped["success"]:
        return jsonify({"success": False, "error": scraped["error"],
                        "site": scraped.get("site",""), "product_name": scraped.get("product_name",""),
                        "reviews": [], "stats": None}), 200
    enriched = []
    for idx, text in enumerate(scraped["reviews"]):
        sentiment, score = analyze_sentiment(text)
        enriched.append({"id":idx,"text":text,"sentiment":sentiment,"score":round(score,4)})
    s = get_sentiment_stats(scraped["reviews"])
    kw = extract_keywords(scraped["reviews"], top_n=25)
    audit("SCRAPE", url[:100])
    return jsonify({"success":True,"url":url,"site":scraped["site"],
                    "original_url":scraped.get("original_url",url),
                    "redirected":scraped.get("redirected",False),
                    "product_name":scraped["product_name"],
                    "reviews":enriched,"stats":s,"keywords":kw,
                    "count":len(enriched),"error":None})


# ── AUDIT + NOTIFICATIONS ─────────────────────────────────────────────────────
@app.route("/audit", methods=["GET"])
@jwt_required()
@roles_required("admin")
def get_audit():
    page     = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))
    q        = AuditLog.query.order_by(AuditLog.created_at.desc())
    result   = paginate(q, page, per_page)
    return jsonify({"logs": [l.to_dict() for l in result["items"]],
                    "pagination": {k:v for k,v in result.items() if k!="items"}})

@app.route("/notifications", methods=["GET"])
@jwt_required()
def get_notifications():
    identity = get_jwt_identity()
    user     = User.query.filter_by(username=identity).first()
    # Admin sees all, others see their own
    if user and user.role == "admin":
        notifs = Notification.query.order_by(Notification.created_at.desc()).limit(50).all()
    else:
        notifs = Notification.query.filter_by(user=identity).order_by(Notification.created_at.desc()).limit(20).all()
    unread = sum(1 for n in notifs if not n.is_read)
    return jsonify({"notifications": [n.to_dict() for n in notifs], "unread": unread})

@app.route("/notifications/<int:nid>/read", methods=["PATCH"])
@jwt_required()
def mark_read(nid):
    n = Notification.query.get_or_404(nid)
    n.is_read = True
    db.session.commit()
    return jsonify({"message": "Marked as read"})

@app.route("/notifications/read-all", methods=["PATCH"])
@jwt_required()
def mark_all_read():
    identity = get_jwt_identity()
    Notification.query.filter_by(user=identity, is_read=False).update({"is_read": True})
    db.session.commit()
    return jsonify({"message": "All marked as read"})

# ── EXPORT URL HELPERS (no auth for direct download links) ────────────────────
def export_url(product, fmt):
    return f"/api/export/{product}/{fmt}"

def export_all_url():
    return "/api/export/all/csv"

# ── ERROR HANDLERS ────────────────────────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Not found"}), 404

@app.errorhandler(429)
def rate_limited(e):
    return jsonify({"error": "Rate limit exceeded. Please slow down."}), 429

@app.errorhandler(500)
def server_error(e):
    logger.error(f"500 error: {e}")
    return jsonify({"error": "Internal server error"}), 500

# ── STARTUP ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5000)

